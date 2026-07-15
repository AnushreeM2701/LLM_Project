import os,re,pdfplumber,pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment
QUESTION_FOLDER="data/aime_questions"
ANSWER_FOLDER="data/aime_answers"
OUTPUT_FILE="data/raw/master_dataset.xlsx"
def extract_text(pdf_path):
    t=""
    with pdfplumber.open(pdf_path) as pdf:
        for p in pdf.pages:
            x=p.extract_text()
            if x:t+=x+"\n"
    return t
def extract_questions(text):
    text=re.sub(r"PAGE\s+\d+","",text)
    qs=[]
    for i in range(1,16):
        s=re.search(rf"\n{i}\.\s","\n"+text)
        if not s: continue
        sp=s.start()
        if i<15:
            e=re.search(rf"\n{i+1}\.\s","\n"+text)
            ep=e.start() if e else len(text)
        else:
            e=re.search(r"Solutions:",text)
            ep=e.start() if e else len(text)
        qs.append(text[sp:ep].strip())
    return qs

def clean_question(question):
    """
    Clean question text extracted from PDF.
    """

    # Remove Answer at the end
    question = re.sub(r"\bAnswer\b", "", question)

    # Remove page numbers
    question = re.sub(r"PAGE\s+\d+", "", question)

    # Replace newlines with spaces
    question = question.replace("\n", " ")

    # Remove multiple spaces
    question = re.sub(r"\s+", " ", question)

    # Remove spaces before punctuation
    question = re.sub(r"\s+([.,;:])", r"\1", question)

    return question.strip()
def extract_answers(text):
    return {int(q):a for q,a in re.findall(r"(\d{1,2})\.\s*([0-9]+)",text) if 1<=int(q)<=15}
def get_year_exam(fn):
    m=re.search(r"(\d{4})(II|I)",fn)
    return (m.group(1),m.group(2)) if m else (None,None)
rows=[]
for qfile in sorted(f for f in os.listdir(QUESTION_FOLDER) if f.endswith(".pdf")):
    year,exam=get_year_exam(qfile)
    qtext=extract_text(os.path.join(QUESTION_FOLDER,qfile))
    qs=extract_questions(qtext)
    afile=qfile.replace("-exam.pdf","-answers.pdf")
    ans={}
    ap=os.path.join(ANSWER_FOLDER,afile)
    if os.path.exists(ap):
        ans=extract_answers(extract_text(ap))
    for i,q in enumerate(qs,1):
        rows.append({"Question ID":f"AIME_{year}_{exam}_P{i:02}","Source":"AIME","Year":int(year),"Exam":exam,"Problem Number":i,"Category":"","Difficulty":"Hard","Include":"T","Question":clean_question(q),"Official Answer":ans.get(i,"")})
df=pd.DataFrame(rows)
os.makedirs("data/raw",exist_ok=True)
with pd.ExcelWriter(OUTPUT_FILE,engine="openpyxl") as w:
    df.to_excel(w,sheet_name="All_Questions",index=False)
wb=load_workbook(OUTPUT_FILE);ws=wb["All_Questions"]
for c in ws[1]: c.font=Font(bold=True)
ws.freeze_panes="A2"
for r in ws.iter_rows():
    for c in r: c.alignment=Alignment(wrap_text=True,vertical="top")
for col,w in {"A":20,"B":10,"C":8,"D":8,"E":16,"F":18,"G":12,"H":12,"I":100,"J":18}.items():
    ws.column_dimensions[col].width=w
wb.save(OUTPUT_FILE)
print("Done!",len(df),"questions")

