"""
Extracts AIME questions + official answers from the source PDFs into
config.AIME_MASTER_PATH.
"""

import os
import re

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from config.config import AIME_MASTER_PATH, RAW_DIR

QUESTION_FOLDER = os.path.join(os.path.dirname(RAW_DIR), "aime_questions")
ANSWER_FOLDER = os.path.join(os.path.dirname(RAW_DIR), "aime_answers")


def extract_text(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text


def extract_questions(text):
    text = re.sub(r"PAGE\s+\d+", "", text)
    questions = []
    for i in range(1, 16):
        start = re.search(rf"\n{i}\.\s", "\n" + text)
        if not start:
            continue
        start_pos = start.start()
        if i < 15:
            end = re.search(rf"\n{i + 1}\.\s", "\n" + text)
            end_pos = end.start() if end else len(text)
        else:
            end = re.search(r"Solutions:", text)
            end_pos = end.start() if end else len(text)
        questions.append(text[start_pos:end_pos].strip())
    return questions


def clean_question(question):
    question = re.sub(r"\bAnswer\b", "", question)
    question = re.sub(r"PAGE\s+\d+", "", question)
    question = question.replace("\n", " ")
    question = re.sub(r"\s+", " ", question)
    question = re.sub(r"\s+([.,;:])", r"\1", question)
    return question.strip()


def extract_answers(text):
    return {
        int(q): a
        for q, a in re.findall(r"(\d{1,2})\.\s*([0-9]+)", text)
        if 1 <= int(q) <= 15
    }


def get_year_exam(filename):
    m = re.search(r"(\d{4})(II|I)", filename)
    return (m.group(1), m.group(2)) if m else (None, None)


def extract_aime_dataset():

    rows = []

    for qfile in sorted(f for f in os.listdir(QUESTION_FOLDER) if f.endswith(".pdf")):

        year, exam = get_year_exam(qfile)
        qtext = extract_text(os.path.join(QUESTION_FOLDER, qfile))
        questions = extract_questions(qtext)

        afile = qfile.replace("-exam.pdf", "-answers.pdf")
        answers = {}
        apath = os.path.join(ANSWER_FOLDER, afile)
        if os.path.exists(apath):
            answers = extract_answers(extract_text(apath))

        for i, q in enumerate(questions, 1):
            rows.append({
                "Question ID": f"AIME_{year}_{exam}_P{i:02}",
                "Source": "AIME",
                "Year": int(year),
                "Exam": exam,
                "Problem Number": i,
                "Category": "",
                "Difficulty": "Hard",
                "Include": "T",
                "Question": clean_question(q),
                "Official Answer": answers.get(i, ""),
            })

    df = pd.DataFrame(rows)

    os.makedirs(os.path.dirname(AIME_MASTER_PATH), exist_ok=True)

    with pd.ExcelWriter(AIME_MASTER_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All_Questions", index=False)

    wb = load_workbook(AIME_MASTER_PATH)
    ws = wb["All_Questions"]
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = {"A": 20, "B": 10, "C": 8, "D": 8, "E": 16, "F": 18, "G": 12, "H": 12, "I": 100, "J": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    wb.save(AIME_MASTER_PATH)

    print(f"Extracted {len(df)} AIME questions -> {AIME_MASTER_PATH}")
    print("NOTE: newly extracted rows need manual Include/Category curation "
          "before they can enter the dataset (see docs/methodology.md).")

    return df


if __name__ == "__main__":
    extract_aime_dataset()
